import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# ==============================
# CONFIGURACIÓN
# ==============================
# Ruta fija para Windows (asegúrate que el archivo exista ahí)
base_path = r"C:\Users\jimgl\Downloads"
file_path = os.path.join(base_path, "RUTA.xlsx")                # Archivo original
output_path = os.path.join(base_path, "RUTA_Procesada.xlsx")    # Archivo de salida

# ==============================
# 1. Cargar archivo y hojas
# ==============================
xls = pd.ExcelFile(file_path)
entregas_df = pd.read_excel(xls, sheet_name='EntregasADomicilio')
orden_df = pd.read_excel(xls, sheet_name='ORDEN')

# ==============================
# 2. Preparar drivers únicos
# ==============================
drivers = orden_df['driver'].unique()

# ==============================
# 3. Abrir workbook con openpyxl para modificar hoja RUTA
# ==============================
wb = load_workbook(file_path)
ruta_ws = wb['RUTA']

# ==============================
# 4. Inicializar posición de pegado en hoja RUTA
# ==============================
current_row = 2

# ==============================
# 5. Procesar cada driver
# ==============================
for driver in drivers:
    print(f"🔄 Procesando driver: {driver}")
    
    # Paso 1. Filtrar ORDEN por driver actual
    orden_driver_df = orden_df[orden_df['driver'] == driver]
    
    # Paso 2. Crear tabla paradas temporal (# y cliente)
    paradas_df = orden_driver_df[['stop_number', 'external_id']].copy()
    paradas_df.columns = ['#', 'cliente']
    
    # Paso 3. Simular BUSCARV con map()
    lookup_dict = dict(zip(paradas_df['cliente'], paradas_df['#']))
    entregas_df['#'] = entregas_df['Num. Cliente'].map(lookup_dict)
    
    # Paso 4. Filtrar eliminando NAs y ordenar por #
    entregas_filtradas = entregas_df.dropna(subset=['#'])
    entregas_filtradas = entregas_filtradas.sort_values(by='#')
    
    # Paso 5. Pegar en hoja RUTA
    # 5.1 Escribir nombre del driver en la celda F(current_row)
    ruta_ws[f'F{current_row}'] = driver
    current_row += 1  # Pasar a la siguiente fila para los datos
    
    # 5.2 Convertir DataFrame a filas y pegarlas
    for r in dataframe_to_rows(entregas_filtradas, index=False, header=True):
        for c_idx, value in enumerate(r, start=2):  # Empezar desde columna B
            ruta_ws.cell(row=current_row, column=c_idx, value=value)
        current_row += 1
    
    # 5.3 Aplicar bordes internos y externos a la tabla pegada
    thin = Side(border_style="thin", color="000000")
    rows = entregas_filtradas.shape[0] + 1  # +1 por encabezado
    cols = entregas_filtradas.shape[1]
    
    for row in ruta_ws.iter_rows(min_row=current_row-rows, max_row=current_row-1,
                                 min_col=2, max_col=1+cols):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Paso 6. Borrar tabla paradas temporal de entregas_df
    entregas_df['#'] = None
    
    # Paso 7. Agregar dos filas de separación antes de la siguiente tabla
    current_row += 2

# ==============================
# 6. Guardar archivo procesado
# ==============================
wb.save(output_path)
print(f"✅ Archivo generado exitosamente en: {output_path}")
